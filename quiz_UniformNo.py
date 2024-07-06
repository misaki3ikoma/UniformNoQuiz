import streamlit as st
import openpyxl
import random
import time


#　＜リーグ／球団リスト定義＞

league = ["日本プロ野球　セ・リーグ",
          "日本プロ野球　パ・リーグ"]
team_name_JPBCentral = ['阪神','巨人','広島','中日','横浜','ヤクルト']
team_name_JPBPacific = ['オリックス','ソフトバンク','日本ハム',
                        '千葉ロッテ','楽天','西武']


#　＜サイドバー＞

#　リーグ／球団の選択

selected_league = st.sidebar.radio("リーグ選択",league)

if selected_league == league[0]:
  selected_team = st.sidebar.selectbox(
      "球団選択",team_name_JPBCentral,
      index = None, placeholder = "球団名を選択してください")
elif selected_league == league[1]:
  selected_team = st.sidebar.selectbox(
      "球団選択",team_name_JPBPacific,
      index = None, placeholder = "球団名を選択してください")
else:
  pass


#　＜メイン画面＞

#　タイトル出力
#　サイドバーよりリーグ／球団を選択したときのみクイズ出力

st.title("UNIFORM NUMBER QUIZ")

if selected_team == None:
    st.write("←←←　サイドバーからリーグと球団を選んで挑戦しよう！")
else:

#　選択したリーグ／球団、Excelデータの登録日時を出力

    st.write(selected_league," 　／　 ",selected_team,"（2024年7月4日時点）")

#　選択リーグのExcel読み込み（ファイル名は変数selected_leagueから設定）
#　選択球団のシート名設定（シート名は変数selected_teamから設定）
#　-----Excelはリーグ毎にファイルを作成、球団毎にシートを作成
#　-----ファイル名、シート名はリスト定義したリーグ／球団と同じ名前にしておく

    from openpyxl import load_workbook

    filename = selected_league + '.xlsx'
    workbook = openpyxl.load_workbook(filename)

    sheet_name = selected_team
    worksheet = workbook[sheet_name]


#　Excelシートの2～80行目から1行をランダム選出してrandom_rowに格納
#　（for文：A列目の背番号データが空白(None)になった時点でbreak）
#　取得データ（A,B列）のデータを変数に格納
#　A列：背番号⇒uniformNo、B列：選手名⇒playerName

    random_row = []
    rows = []
    for row in worksheet.iter_rows(min_row=2, max_row=80, values_only=True):
        values = [cell for cell in row]
        if values[0] is None :
            break
        else:
            rows.append(row)
    
    random_row = random.choice(rows)

    uniformNo = random_row[0]
    playerName = random_row[1]

#　デバック用出力コード
#    index = rows.index(random_row)
#    length = len(rows)
#    st.write(index)
#    st.write(length)


#　出題形式の選択
#　背番号⇒選手名、もしくは、選手名⇒背番号
#　変数Question、Answerに選択した形式で格納

    quiz_format = ["背番号⇒選手名","選手名⇒背番号"]
    selected_Qformat = st.radio("出題形式を選択",quiz_format,horizontal=True)

    if selected_Qformat == quiz_format[0]:
        Question = uniformNo
        Answer = playerName
    elif selected_Qformat == quiz_format[1]:
        Question = playerName
        Answer = uniformNo
    else:
        pass
       

#　出題ボタン、解答ボタンのアクション

    if st.button("問題") :
        time.sleep(1) 
        st.markdown(f'<span style="font-size:24px">{Question}</span>',
                       unsafe_allow_html=True)
        with st.expander("答えは・・・"):
            st.markdown(f'<span style="color:red;font-size:24px">{Answer}</span>',
                          unsafe_allow_html=True)

#　次の問題ボタンのアクション

        if st.button("次の問題へ"):
            placeholder = st.empty()
            placeholder.empty()
        else:
            pass
    else:
        pass       
